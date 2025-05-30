from curses.ascii import isdigit
import decimal
from pickletools import read_stringnl_noescape_pair
from openpyxl import load_workbook
import pytexit
import re

# this import sys bad, hack for easy stdout redirect to file
import sys
sys.stdout.reconfigure(encoding='utf-8')
filename = "aue_lab7.xlsx"
workbook = load_workbook(filename=filename )
workbook_data = load_workbook(filename=filename, data_only=True )
sheet = workbook.active
datasheet = workbook_data.active


def find_cell(s:str):
    # returns if string begins with cell eg. "A5", "BC52"
    # if doesnt find, return empty string
    
    find_cell_reg = r"^[A-Z]+\d+"
    match = re.search(find_cell_reg,s)
    if match:
        return str(match.group(0))
    else:
        return ''

# this is necessary
from math import sqrt

def clean_formula(val):
    # replaces excel to python lang
    val= val.replace('^','**')
    val = val.replace("SQRT" , 'sqrt')
    val = val.replace("-0" , '-')
    val = val.replace("PI()" , '3.14')
    return val

def build_equation(cell_value):
    # finds the equation and values in the cell and builds a string
    s = ''
    i = 1
    if not cell_value:
        return "0"
    if isinstance(cell_value,(int,float)) :
        return str(cell_value)

    while i != len(cell_value):
        newcell = find_cell(cell_value[i:])
        if newcell:
            values = clean_formula(build_equation(sheet[newcell].value))
            s += str(eval(values))
            i += len(newcell)-1
        else:
            s+=cell_value[i]
        i+=1
    return clean_formula(s)

        

def create_latex(cell):
    # creates latex 
    equation =build_equation(cell)
    latex = pytexit.py2tex(equation, print_latex=False, print_formula=False)
    return latex

def shorten_small_float(n : str, round_to):
    # run for float<1
    output = n[:2]
    count =0
    counting = False
    # cut out "0."" wtih [2:]
    for ch in n[2:]:

        output+=ch
        if int(ch)>0:
            # found nonzero digit
            counting=True
        if counting:
            count+=1
        if count == round_to:
            return output
    return output

def reduce_floats(s: str, round_to=2):
    find_floats = r"\d+[.]\d+"
    #teststr = repr("Tl1= = $$9.15\times{10}^{-12} \left(4600+292.7351842831629\times{10}^6\right)$$ = 0.0026785690261909405")

    def matchfloat(matchobj):
        round_to = 2
        match = float(matchobj.group(0))
        if match<1:
            # so that you don't turn 0.004 => 0.0
            return shorten_small_float(str(match), round_to)

        return str(round(match, round_to))
    
    value_short = re.sub(find_floats, matchfloat ,s)
    return value_short


if __name__ == "__main__":
    cells = []

    # set cell search scope
    start_cell = 22
    end_cell   = 40
    value_letter = 'u'


    search_col = int(ord(value_letter))-int(ord('a'))+1
    for i in range(start_cell,end_cell):
        cell_name = sheet.cell(row=i, column=search_col-1).value
        #print(sheet.cell(row=i, column=search_col).value[1:])
        latex = create_latex(sheet.cell(row=i, column=search_col).value)
        value = datasheet.cell(row=i, column=search_col).value

        output = reduce_floats(f'{cell_name} = {latex} = {value}',2)
        #print(f'{cell_name} = {latex} = {value}')
        print(output)
        #print((sheet.cell(row=i, column=search_col).value))
